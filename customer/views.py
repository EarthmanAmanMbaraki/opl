
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.status import (
	HTTP_400_BAD_REQUEST,
	HTTP_201_CREATED,
	HTTP_404_NOT_FOUND,
	)
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill,  Font, Alignment, Protection
from rest_framework.views import APIView
from rest_framework.generics import (
	# CreateAPIView,
	# ListAPIView,
	ListCreateAPIView,
	# RetrieveAPIView,
	# RetrieveUpdateAPIView,
	# UpdateAPIView,
	)

from . models import (
	Customer,
    Truck,
)
from depot.models import Depot, Product
from . serializers import (
	CustomerCreateSer,
    CustomerListSer,
    TruckCreateSer,
    TruckListSer,
)

class CustomerCreate(ListCreateAPIView):
    serializer_class = CustomerCreateSer
    queryset = Customer.objects.all()

    def get(self, request, *args, **kwargs):
        customers = Customer.objects.all()
        serializer = CustomerListSer(customers, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        serializer = CustomerCreateSer(data=request.data)

        if serializer.is_valid():
            customer = serializer.save()
            depot = customer.depot
            serializer = CustomerListSer(customer)
            create_excel(depot)
            return Response(serializer.data)

class TruckCreate(ListCreateAPIView):
    serializer_class = TruckCreateSer
    queryset = Truck.objects.all()

    def get(self, request, *args, **kwargs):
        trucks = Truck.objects.all()
        serializer = TruckListSer(data=trucks, many=True)
        return Response(serializer.data)
    
    def post(self, request, *args, **kwargs):
        serializer = TruckCreateSer(data=request.data)

        if serializer.is_valid():
            truck = serializer.save()
            serializer = TruckListSer(truck)
            return Response(serializer.data)

def customers_formula(depot):
    customers = Customer.objects.filter(depot=depot).order_by("name")
    string = ""
    for customer in customers:
        string += f"{customer.name},"
    return string

def products_formula(depot):
    products = Product.objects.filter(depot=depot).order_by("name")
    string = ""
    for product in products:
        string += f"{product.name},"
    return string

def create_excel(depot):
    wb = Workbook()
    font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

    columns = ["DATE", "PRODUCT", "CUSTOMER", "ORDER NO", "ENTRY NO", "VOL OBS", "VOL 20", "SELLING PRICE"]
    labels = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    customers = customers_formula(depot)
    products = products_formula(depot)
    dv = DataValidation(type="list", formula1=f'"{customers}"', allow_blank=False)
    dv2 = DataValidation(type="list", formula1=f'"{products}"', allow_blank=False)
    
    sheet = wb.active
    
    sheet["A1"].value = "ELDORET DEPOT"
    sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].fill =  PatternFill("solid", start_color="00CCFFCC")
    sheet["A1"].protection = Protection(locked=True)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
   
    for idx, column in enumerate(columns):
        sheet[f"{labels[idx]}{2}"] = column
        
        sheet[f"{labels[idx]}{2}"].font = font
        sheet[f"{labels[idx]}{2}"].fill = PatternFill("solid", start_color="5cb800")

        sheet.column_dimensions[labels[idx]].width = 20.0
    sheet.row_dimensions[1].height = 27.0
    dv.add("C3:C100000")
    dv2.add("B3:b100000")
    sheet.add_data_validation(dv)
    sheet.add_data_validation(dv2)
    wb.save("DailyReportTemplate.xlsx")
