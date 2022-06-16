import datetime
from django.http import HttpResponse
from django.core.cache import cache
from rest_framework.response import Response
from rest_framework.generics import ListAPIView, RetrieveAPIView
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Protection
from openpyxl.utils import get_column_letter
from customer.models import Customer
from depot.models import Depot
from product.models import Product

from depot.serializers import (
    RetrieveDepotSer,
    DepotMonthSer,
    DepotProductMonthSer,
    DepotSeriesSer,
    DepotProductSeriesSer,
    DepotCustomerMonthSer,
)


class DepotListView(ListAPIView):
    serializer_class = RetrieveDepotSer
    queryset = Depot.objects.all()

    def get(self, request, *args, **kwargs):
        e = cache.get("depot", None)
        if not e:
            depots = Depot.objects.all()
            serializer = RetrieveDepotSer(depots, many=True)
            cache.set("depot", serializer.data)
            return Response(serializer.data)
        return Response(e)


class DepotMonthView(ListAPIView):
    serializer_class = DepotMonthSer
    queryset = Depot.objects.prefetch_related("sale_set")

    def get_serializer_context(self):
        year = self.request.GET.get("year", None)
        return {"year": year}

    def get(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        e = cache.get("depot-month-{}".format(context["year"]), None)
        if not e:

            serializer = DepotMonthSer(self.get_queryset(), many=True, context=context)
            cache.set("depot-month-{}".format(context["year"]), serializer.data)
            return Response(serializer.data)
        return Response(e)


class DepotProductMonthView(ListAPIView):
    serializer_class = DepotProductMonthSer
    queryset = Depot.objects.prefetch_related("sale_set")

    def get_serializer_context(self):
        year = self.request.GET.get("year", None)
        return {"year": year}

    def get(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        e = cache.get("depot-product-month-{}".format(context["year"]))
        if not e:
            serializer = DepotProductMonthSer(
                self.get_queryset(), many=True, context=context
            )
            cache.set("depot-product-month-{}".format(context["year"]), serializer.data)
            return Response(serializer.data)
        return Response(e)


class DepotSeriesView(RetrieveAPIView):
    serializer_class = DepotSeriesSer
    queryset = Depot.objects.prefetch_related("sale_set")

    def get_serializer_context(self):
        start_date = self.request.GET.get("start_date", None)
        end_date = self.request.GET.get("end_date", None)
        if start_date and end_date:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        return {"start_date": start_date, "end_date": end_date}

    def get(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        e = cache.get(
            "depot-series-{}-{}-{}".format(
                self.get_object().id, context["start_date"], context["end_date"]
            ),
            None,
        )
        if not e:
            serializer = DepotSeriesSer(self.get_object(), context=context)
            cache.set(
                "depot-customer-{}-{}-{}".format(
                    self.get_object().id, context["start_date"], context["end_date"]
                ),
                serializer.data,
            )
            return Response(serializer.data)
        return Response(e)


class DepotProductSeriesView(RetrieveAPIView):
    serializer_class = DepotProductSeriesSer
    queryset = Depot.objects.prefetch_related("sale_set")

    def get_serializer_context(self):
        start_date = self.request.GET.get("start_date", None)
        end_date = self.request.GET.get("end_date", None)
        if start_date and end_date:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        return {"start_date": start_date, "end_date": end_date}

    def get(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        e = cache.get(
            "depot-product-series-{}-{}-{}".format(
                self.get_object().id, context["start_date"], context["end_date"]
            ),
            None,
        )
        if not e:
            serializer = DepotProductSeriesSer(self.get_object(), context=context)
            cache.set(
                "depot-customer-{}-{}-{}".format(
                    self.get_object().id, context["start_date"], context["end_date"]
                ),
                serializer.data,
            )
            return Response(serializer.data)
        return Response(e)


class DepotCustomerMonthView(RetrieveAPIView):
    serializer_class = DepotCustomerMonthSer
    queryset = Depot.objects.prefetch_related("sale_set")

    def get_serializer_context(self):
        year = self.request.GET.get("year", None)
        month = int(self.request.GET.get("month", None))
        return {"year": year, "month": month}

    def get(self, request, *args, **kwargs):
        context = self.get_serializer_context()
        e = cache.get(
            "depot-customer-{}-{}-{}".format(
                self.get_object().id, context["year"], context["month"]
            ),
            None,
        )
        if not e:
            serializer = DepotCustomerMonthSer(self.get_object(), context=context)
            cache.set(
                "depot-customer-{}-{}-{}".format(
                    self.get_object().id, context["year"], context["month"]
                ),
                serializer.data,
            )
            return Response(serializer.data)
        return Response(e)


def customers_formula(depot):
    # customers = depot.depotcustomer_set.all().order_by("customer__name")
    customers = Customer.objects.all()
    return customers


def products_formula():
    products = Product.objects.all()
    string = ""
    for product in products:
        string += f"{product.name},"
    return string


def create_excel(depot):
    wb = Workbook()
    font = Font(
        name="Calibri",
        size=11,
        bold=True,
        italic=False,
        vertAlign=None,
        underline="none",
        strike=False,
        color="FF000000",
    )

    columns = [
        "DATE",
        "PRODUCT",
        "CUSTOMER",
        "TRUCK",
        "ORDER NO",
        "LPO_NO",
        "ENTRY NO",
        "VOL OBS",
        "VOL 20",
        "SELLING PRICE",
        "PAYMENT",
    ]
    customers = customers_formula(depot)
    products = products_formula()
    sheet = wb.active
    ws2 = wb.create_sheet(title="customers")
    ws3 = wb.create_sheet(title="trucks")
    letter = None
    for idx, customer in enumerate(customers):
        ws2[f"A{idx+1}"] = customer.name
        letter = get_column_letter(idx + 1)
        ws3[f"{letter}1"] = customer.name
        for index, truck in enumerate(customer.truck_set.all()):
            ws3[f"{letter}{index+2}"] = truck.plate_no
    dv = DataValidation(
        type="list",
        formula1="=OFFSET('customers'!$A$1,0,0,COUNTA('customers'!$A:$A) - 0,1)",
        allow_blank=False,
    )
    dv2 = DataValidation(type="list", formula1=f'"{products}"', allow_blank=False)
    dv3 = DataValidation(
        type="list",
        formula1=f"=OFFSET(trucks!$A$1,1, MATCH($C3, trucks!$A$1:${letter}$1,0)-1,COUNTA(OFFSET(trucks!$A$1,1, MATCH($C3, trucks!$A$1:${letter}$1,0)-1,200,1)),1)",
        allow_blank=True,
    )
    dv4 = DataValidation(type="list", formula1=f'"YES, NO"', allow_blank=False)

    sheet["A1"].value = depot.name
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].fill = PatternFill("solid", start_color="00CCFFCC")
    sheet["A1"].protection = Protection(locked=True)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    for idx, column in enumerate(columns):
        l = get_column_letter(idx + 1)
        sheet[f"{l}{2}"] = column

        sheet[f"{l}{2}"].font = font
        sheet[f"{l}{2}"].fill = PatternFill("solid", start_color="5cb800")

        sheet.column_dimensions[l].width = 20.0
    sheet.row_dimensions[1].height = 27.0
    dv.add("C3:C100000")
    dv3.add("D3:D100000")
    dv2.add("B3:b100000")
    dv4.add("K3:K100000")
    sheet.add_data_validation(dv)
    sheet.add_data_validation(dv2)
    sheet.add_data_validation(dv3)
    sheet.add_data_validation(dv4)
    wb.save(f"DailyReportTemplate{depot.id}.xlsx")


def download(request, depot_id):
    create_excel(Depot.objects.get(pk=int(depot_id)))
    with open(f"DailyReportTemplate{depot_id}.xlsx", "rb") as fh:
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = (
            "inline; filename=" + "DailyReportTemplate.xlsx"
        )
        return response
    # return HttpResponse("ok")
